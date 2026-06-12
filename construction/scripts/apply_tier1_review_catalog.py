#!/usr/bin/env python3
"""Apply the expert *Tier-1* review pass to the catalog candidate pool.

This is the catalog **Tier-1 review** apply step — distinct from the Core-4
evidence audit applied by ``scripts/apply_audit_decisions.py``. The expert
returned a worksheet (``audit_worksheet_catalog_REVIEW_tier1.xlsx``) in which
they:

  * made a scope decision per row (``action`` ∈ {confirm, drop}); and
  * rewrote / enriched the Tier-1 product description (``tier1_expert``,
    ``tier1_detailed``, ``t1_family``, ``tier1_rep_part``, ``tier1_grounding``,
    ``tier1_expert_source``, ``t1_confidence``, ``tier1_description_normalized``).

They did NOT fill the Core-4 evidence columns (``expert_hs6``,
``confidence_tier``, ``cited_evidence_ids``, ``adjudication_status``) — those
belong to a separate, later evidence-grounding pass. So this script does not
flip ``label_source`` to the released enum; surviving records stay
``catalog_expert_validated_pending_reaudit`` and feed a second (Core-4)
worksheet.

What this script does:

  1. Reads the review file (``.xlsx`` Review sheet, or ``.csv``), keyed by
     ``record_id`` (== candidate-pool ``id``).
  2. Drops every ``action == "drop"`` record from the catalog pool.
  3. For every ``action == "confirm"`` record, folds the enriched Tier-1
     description into the record:
       * ``tier1_description`` ← ``tier1_expert`` (canonical; falls back to
         ``tier1_detailed`` then the original text if blank).
       * Original text + all enrichment fields preserved under a
         ``tier1_review`` provenance block so the rewrite is fully reversible.
       * ``tier1_source`` set to ``"catalog_expert_tier1"`` to mark that the
         Tier-1 text is expert-authored, not the raw catalog string.
  4. Leaves ``hs6_label`` / ``candidate_set`` untouched (the worksheet had
     zero ``action=change`` rows; confirm means "HS6 is right as-is, pending
     Core-4 evidence").
  5. Writes the surviving pool + a review report.

HS6 labels are unchanged, so ``candidate_set`` and tier2 stay valid.
tier2_minimal is intentionally NOT regenerated from the
enriched Tier-1 text — that would be a separate degradation re-run.

Idempotent given the same inputs. Deterministic ordering (by hs6_label + id).

Outputs:
  * ``release/working/data/_candidate_pool_catalog_tier1.jsonl`` — survivors
    with folded Tier-1 enrichment (input to the Core-4 worksheet).
  * ``release/working/data/_catalog_tier1_review_report.json`` — drop/confirm
    counts, per-HS6 breakdown, enrichment + grounding coverage, dropped ids.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

ROOT = Path(__file__).resolve().parents[1]
WORKING_DATA = ROOT / "release" / "working" / "data"

DEFAULT_REVIEW_FILE = (
    Path.home()
    / "Downloads"
    / "semihs_review_2026-05-26"
    / "catalog"
    / "audit_worksheet_catalog_REVIEW_tier1.xlsx"
)
DEFAULT_POOL_IN = WORKING_DATA / "_candidate_pool_catalog.jsonl"
DEFAULT_POOL_OUT = WORKING_DATA / "_candidate_pool_catalog_tier1.jsonl"
DEFAULT_REPORT_OUT = WORKING_DATA / "_catalog_tier1_review_report.json"
DEFAULT_SHEET = "Review"

PENDING_LABEL = "catalog_expert_validated_pending_reaudit"

# Worksheet columns we read. record_id links to pool 'id'.
_ID_COL = "record_id"
_ACTION_COL = "action"
# Tier-1 enrichment columns (expert-authored).
_ENRICH_COLS = (
    "tier1_description_normalized",
    "t1_family",
    "t1_confidence",
    "t1_flags",
    "t1_validation",
    "tier1_expert",
    "tier1_expert_source",
    "tier1_detailed",
    "tier1_rep_part",
    "tier1_grounding",
)


def _rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


# ---------------------------------------------------------------------------
# Review-file loading (xlsx or csv)
# ---------------------------------------------------------------------------

def _read_review_rows(path: Path, sheet: str) -> List[Dict[str, str]]:
    """Return review rows as a list of dicts keyed by column header.

    Supports ``.xlsx`` (reads the given sheet via openpyxl) and ``.csv``.
    """
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        try:
            import openpyxl  # type: ignore
        except ImportError as exc:  # pragma: no cover
            raise SystemExit(
                "openpyxl is required to read .xlsx review files. "
                "Install it (pip install openpyxl) or export the Review sheet "
                "to CSV and pass that instead."
            ) from exc
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            raise SystemExit(
                f"sheet {sheet!r} not found in {path.name}; "
                f"available: {wb.sheetnames}"
            )
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(it)]
        rows: List[Dict[str, str]] = []
        for raw in it:
            if raw is None:
                continue
            rec = {
                header[i]: ("" if v is None else str(v))
                for i, v in enumerate(raw)
                if i < len(header) and header[i]
            }
            # Skip fully empty trailing rows.
            if any(v.strip() for v in rec.values()):
                rows.append(rec)
        return rows
    # CSV fallback.
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [dict(r) for r in csv.DictReader(f)]


def _read_jsonl(path: Path) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            out.append(json.loads(line))
    return out


def _write_jsonl(path: Path, records: Sequence[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, sort_keys=True, ensure_ascii=False) + "\n")


# ---------------------------------------------------------------------------
# Decision indexing
# ---------------------------------------------------------------------------

def _index_decisions(rows: Sequence[Mapping[str, str]]) -> Dict[str, Dict[str, str]]:
    """Map record_id → decision dict. Errors on blank or duplicate ids."""
    out: Dict[str, Dict[str, str]] = {}
    for i, row in enumerate(rows, start=2):  # +2 ≈ spreadsheet line (header=1)
        rid = (row.get(_ID_COL) or "").strip()
        if not rid:
            raise SystemExit(f"review row {i}: blank {_ID_COL}")
        if rid in out:
            raise SystemExit(f"review row {i}: duplicate {_ID_COL} {rid!r}")
        out[rid] = dict(row)
    return out


def _canonical_tier1(decision: Mapping[str, str], original: str) -> Tuple[str, str]:
    """Return (canonical_tier1_text, which_field). Prefer expert prose, then
    detailed, then keep the original."""
    expert = (decision.get("tier1_expert") or "").strip()
    if expert:
        return expert, "tier1_expert"
    detailed = (decision.get("tier1_detailed") or "").strip()
    if detailed:
        return detailed, "tier1_detailed"
    return original, "original_unchanged"


def _build_tier1_review_block(
    decision: Mapping[str, str], original_tier1: str, canonical_field: str
) -> Dict[str, Any]:
    """Provenance block capturing the full expert Tier-1 enrichment, so the
    rewrite is reversible and downstream can audit grounding."""
    block: Dict[str, Any] = {
        "reviewed_action": "confirm",
        "original_tier1_description": original_tier1,
        "canonical_source_field": canonical_field,
        "review_pass": "tier1",
    }
    for col in _ENRICH_COLS:
        val = (decision.get(col) or "").strip()
        if val:
            block[col] = val
    return block


# ---------------------------------------------------------------------------
# Apply
# ---------------------------------------------------------------------------

def apply_tier1_review(
    pool: Sequence[Mapping[str, Any]],
    decisions: Mapping[str, Mapping[str, str]],
    *,
    drop_unreviewed: bool = False,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Return (survivors, report-fragment).

    ``drop_unreviewed``: when False (default, catalog full-pool case) pool
    records with no matching review decision are kept untouched. When True
    (carryover-subset case, where the review covers only the catalog-origin
    records) they are dropped from the output — so feeding the full 133-record
    carryover pool yields only the reviewed catalog-origin survivors and leaves
    the BOL-origin carryover for its own pass.
    """
    survivors: List[Dict[str, Any]] = []
    dropped_ids: List[str] = []
    confirmed_ids: List[str] = []
    missing_decision: List[str] = []
    unknown_action: List[Tuple[str, str]] = []
    fallback_canonical: List[str] = []

    per_hs6: Dict[str, Counter] = defaultdict(Counter)
    grounding_counts: Counter = Counter()
    family_counts: Counter = Counter()
    t1_confidence_counts: Counter = Counter()

    for rec in pool:
        rid = str(rec.get("id") or "")
        decision = decisions.get(rid)
        hs6 = str(rec.get("hs6_label") or "")
        if decision is None:
            missing_decision.append(rid)
            if not drop_unreviewed:
                # Conservative: keep undecided records untouched.
                survivors.append(dict(rec))
            continue

        action = (decision.get(_ACTION_COL) or "").strip().lower()
        per_hs6[hs6][action or "blank"] += 1

        if action == "drop":
            dropped_ids.append(rid)
            continue
        if action != "confirm":
            unknown_action.append((rid, action))
            # Keep, but flag — do not silently drop.
            survivors.append(dict(rec))
            continue

        # confirm → fold enrichment.
        new_rec = dict(rec)
        original_tier1 = str(rec.get("tier1_description") or "")
        canonical, field = _canonical_tier1(decision, original_tier1)
        if field == "original_unchanged":
            fallback_canonical.append(rid)
        new_rec["tier1_description"] = canonical
        # tier1_source stays within the schema enum {BOL, catalog}; the fact
        # that the Tier-1 text is expert-authored is recorded in the
        # tier1_review provenance block (canonical_source_field + review_pass).
        new_rec["tier1_source"] = "catalog"
        new_rec["tier1_review"] = _build_tier1_review_block(
            decision, original_tier1, field
        )
        # label_source stays pending — Core-4 evidence audit still owed.
        new_rec["label_source"] = PENDING_LABEL
        survivors.append(new_rec)
        confirmed_ids.append(rid)

        grounding_counts[(decision.get("tier1_grounding") or "unknown").strip() or "unknown"] += 1
        family_counts[(decision.get("t1_family") or "unknown").strip() or "unknown"] += 1
        t1_confidence_counts[(decision.get("t1_confidence") or "unknown").strip() or "unknown"] += 1

    # Deterministic ordering (hs6 then id) — ids/frozen_ids preserved as-is
    # for traceability back to the Tier-1 worksheet.
    survivors.sort(key=lambda r: (str(r.get("hs6_label") or ""), str(r.get("id") or "")))

    report = {
        "counts": {
            "pool_in": len(pool),
            "confirmed": len(confirmed_ids),
            "dropped": len(dropped_ids),
            "survivors_out": len(survivors),
            "missing_decision": len(missing_decision),
            "unknown_action": len(unknown_action),
            "canonical_fallback_to_original": len(fallback_canonical),
        },
        "per_hs6": {h6: dict(sorted(c.items())) for h6, c in sorted(per_hs6.items())},
        "enrichment_coverage": {
            "by_grounding": dict(sorted(grounding_counts.items())),
            "by_t1_confidence": dict(sorted(t1_confidence_counts.items())),
            "by_family": dict(family_counts.most_common()),
        },
        "dropped_ids": sorted(dropped_ids),
        "missing_decision_ids": sorted(missing_decision),
        "unknown_action_rows": sorted(unknown_action),
        "canonical_fallback_ids": sorted(fallback_canonical),
    }
    return survivors, report


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(argv: Optional[Sequence[str]] = None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--review-file", type=Path, default=DEFAULT_REVIEW_FILE,
                   help=f"Expert Tier-1 review file (.xlsx or .csv). "
                        f"Default: {DEFAULT_REVIEW_FILE}")
    p.add_argument("--sheet", default=DEFAULT_SHEET,
                   help=f"Worksheet/tab name for .xlsx (default: {DEFAULT_SHEET}).")
    p.add_argument("--candidate-pool", type=Path, default=DEFAULT_POOL_IN,
                   help=f"Catalog candidate pool (default: {_rel(DEFAULT_POOL_IN)}).")
    p.add_argument("--output", type=Path, default=DEFAULT_POOL_OUT,
                   help=f"Output JSONL (default: {_rel(DEFAULT_POOL_OUT)}).")
    p.add_argument("--report-out", type=Path, default=DEFAULT_REPORT_OUT,
                   help=f"Report JSON (default: {_rel(DEFAULT_REPORT_OUT)}).")
    p.add_argument("--drop-unreviewed", action="store_true",
                   help="Drop pool records that have no matching review row "
                        "(instead of keeping them). Use when the review covers "
                        "only a subset of the pool — e.g. the catalog-origin "
                        "subset of the carryover pool.")
    p.add_argument("--dry-run", action="store_true",
                   help="Compute and print counts; do not write outputs.")
    args = p.parse_args(argv)

    if not args.review_file.exists():
        raise SystemExit(f"review file not found: {args.review_file}")
    if not args.candidate_pool.exists():
        raise SystemExit(f"candidate pool not found: {args.candidate_pool}")

    review_rows = _read_review_rows(args.review_file, args.sheet)
    print(f"read {len(review_rows)} review rows from {args.review_file.name}")
    decisions = _index_decisions(review_rows)

    pool = _read_jsonl(args.candidate_pool)
    print(f"read {len(pool)} catalog pool records from {_rel(args.candidate_pool)}")

    survivors, report = apply_tier1_review(
        pool, decisions, drop_unreviewed=args.drop_unreviewed
    )

    c = report["counts"]
    print(f"  confirmed (kept): {c['confirmed']}")
    print(f"  dropped:          {c['dropped']}")
    print(f"  survivors out:    {c['survivors_out']}")
    if c["missing_decision"]:
        disp = "dropped" if args.drop_unreviewed else "kept untouched"
        print(f"  ⚠ {c['missing_decision']} pool records had NO matching review row ({disp})")
    if c["unknown_action"]:
        print(f"  ⚠ {c['unknown_action']} rows had an unrecognized action (kept untouched)")
    if c["canonical_fallback_to_original"]:
        print(f"  ⚠ {c['canonical_fallback_to_original']} confirms had blank expert text — kept original tier1")
    print(f"  grounding: {report['enrichment_coverage']['by_grounding']}")

    if args.dry_run:
        print("\nDRY RUN — not writing outputs.")
        return 0

    _write_jsonl(args.output, survivors)
    args.report_out.parent.mkdir(parents=True, exist_ok=True)
    args.report_out.write_text(
        json.dumps(report, indent=2, sort_keys=True, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print(f"\nwrote {_rel(args.output)} ({len(survivors)} records)")
    print(f"wrote {_rel(args.report_out)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
